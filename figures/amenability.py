#!/usr/bin/env python
"""
Real microbial metabolomics — amenability probes across organisms.

Generalises the E. coli amenability analysis to the three organisms measured by
Link et al. 2015 (nmeth.3584): E. coli, B. subtilis, Y. lipolytica. Same four
probes per organism, so the figures are directly comparable and answer: does
the "low-rank + network-mappable but noisy dc/dt" reading generalise, or is it
organism-specific?

Per organism it recapitulates figures/metabolism/amenability_<org>.png:
  H5 low-rank      : SVD of the (ion x time) matrix -> rank@90/99%.
  H6 predictability: per-ion lag-1 autocorrelation (smooth signal vs noise).
  (c) example traces: highest-variance ions (z-scored).
  H7 network cover : measured KEGG ions mapped onto a reference bacterial GEM
                     (iJO1366) as a proxy for "core/conserved metabolite that a
                     genome-scale model would contain". Organism-specific GEMs
                     (iYO844, iYali4) would refine B. subtilis / Y. lipolytica.

Inputs (vendored): papers/nmeth3584/...MOESM197...xlsx ; papers/iJO1366.xml,
papers/e_coli_core.xml.  Run: python figures/amenability.py
"""
import os, re
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib as _mpl
_mpl.rcParams["axes.spines.top"] = False; _mpl.rcParams["axes.spines.right"] = False  # bare x/y axes
import matplotlib.pyplot as plt
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "papers/nmeth3584/41592_2015_BFnmeth3584_MOESM197_ESM.xlsx")
OUTDIR = os.path.join(ROOT, "figures/metabolism")
os.makedirs(OUTDIR, exist_ok=True)

# plot style: large fonts overall; ONE bold organism label (a/b/c/d) per figure,
# top-left, no panel titles and no organism name on the figure.
plt.rcParams.update({
    "font.size": 16, "axes.labelsize": 18, "axes.titlesize": 16,
    "xtick.labelsize": 14, "ytick.labelsize": 14, "legend.fontsize": 14,
})


def figure_tag(fig, letter):
    fig.text(0.012, 0.985, letter, fontsize=30, fontweight="bold",
             va="top", ha="left")

ORGANISMS = [
    ("E. coli",      "ecoli", "Ecoli1", "Annotation Ecoli"),
    ("B. subtilis",  "bsubt", "Bsubt1", "Annotation Bsubt"),
    ("Y. lipolytica","ylipo", "Ylipo1", "Annotation Ylipo"),
    ("Hybrid",       "hybr",  "Hybr1",  "Annotation Hybr"),
]


def load_organism(wb, data_sheet, ann_sheet):
    ws = wb[data_sheet]
    rows = list(ws.iter_rows(values_only=True))
    tvec = np.array([v for v in rows[1][1:] if v is not None], dtype=float)
    T = len(tvec)
    data = []
    for r in rows[2:]:
        if r[0] is None:
            continue
        vals = [np.nan if v is None else float(v) for v in r[1:1 + T]]
        if len(vals) == T:
            data.append(vals)
    C = np.array(data, dtype=float)
    ann = wb[ann_sheet]
    idx2kegg = {}
    for row in ann.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        idx2kegg.setdefault(int(row[0]), set())
        if row[2]:
            idx2kegg[int(row[0])].add(str(row[2]).strip())
    return tvec, C, idx2kegg


def model_kegg_ids(xml_path):
    if not os.path.exists(xml_path):
        return None
    txt = open(xml_path, "r", errors="ignore").read()
    ids = set(re.findall(r"kegg\.compound[/:]?(C\d{5})", txt))
    if not ids:
        ids = set(re.findall(r"\bC\d{5}\b", txt))
    return ids


def zscore_rows(C):
    mu = np.nanmean(C, axis=1, keepdims=True)
    sd = np.nanstd(C, axis=1, keepdims=True); sd[sd == 0] = 1.0
    return np.nan_to_num((C - mu) / sd)


def svd_rank(Z):
    s = np.linalg.svd(Z, compute_uv=False)
    cum = np.cumsum(s ** 2) / (s ** 2).sum()
    return cum, int(np.searchsorted(cum, .90) + 1), int(np.searchsorted(cum, .99) + 1)


def lag1_autocorr(C):
    out = []
    for row in C:
        x = row[~np.isnan(row)]
        if len(x) < 5 or np.std(x) == 0:
            continue
        x = x - x.mean(); d = np.sum(x * x)
        out.append(np.sum(x[1:] * x[:-1]) / d if d > 0 else 0.0)
    return np.array(out)


def run_one(org, slug, tag, t, C, idx2kegg, ref_kegg, ref_name):
    N, T = C.shape
    Z = zscore_rows(C)
    cum, r90, r99 = svd_rank(Z)
    ac = lag1_autocorr(C)
    ions = sorted(idx2kegg)
    n_ion = len(ions)
    n_kegg = sum(1 for i in ions if idx2kegg[i])
    n_unambig = sum(1 for i in ions if len(idx2kegg[i]) == 1)
    n_ref = sum(1 for i in ions if ref_kegg and (idx2kegg[i] & ref_kegg))

    fig, ax = plt.subplots(2, 2, figsize=(12, 8))
    # (a) SVD low-rank
    ax[0, 0].plot(np.arange(1, len(cum) + 1), cum, "o-", ms=3, color="#2c3e50")
    ax[0, 0].axhline(.99, ls="--", c="#e74c3c", alpha=.6); ax[0, 0].axhline(.90, ls="--", c="#f39c12", alpha=.6)
    ax[0, 0].axvline(r99, ls=":", c="#e74c3c", alpha=.6)
    ax[0, 0].set_xlabel("singular component"); ax[0, 0].set_ylabel("cumulative variance")
    ax[0, 0].set_xlim(0, min(60, len(cum)))
    # (b) autocorrelation
    ax[0, 1].hist(ac, bins=30, color="#3498db", edgecolor="white")
    ax[0, 1].axvline(np.median(ac), c="#e74c3c", label=f"median = {np.median(ac):.2f}")
    ax[0, 1].set_xlabel("lag-1 autocorrelation"); ax[0, 1].set_ylabel("number of ions"); ax[0, 1].legend()
    # (c) example traces
    for j in np.argsort(-np.nanvar(C, axis=1))[:6]:
        ax[1, 0].plot(t, Z[j], lw=1.3)
    ax[1, 0].set_xlabel("time"); ax[1, 0].set_ylabel("z(intensity)")
    # (d) network coverage
    labels = ["ions", "with\nKEGG", "unambig.", f"in\n{ref_name}"]
    vals = [n_ion, n_kegg, n_unambig, n_ref]
    for b, v in zip(ax[1, 1].bar(labels, vals,
                    color=["#95a5a6", "#3498db", "#2980b9", "#16a085"]), vals):
        ax[1, 1].text(b.get_x() + b.get_width() / 2, v + 1, str(v), ha="center", fontsize=14)
    ax[1, 1].set_ylabel("number of ions")
    fig.tight_layout(rect=[0.02, 0, 1, 1])
    figure_tag(fig, tag)
    out = os.path.join(OUTDIR, f"amenability_{slug}.png")
    fig.savefig(out, dpi=130); plt.close(fig)
    verdict = dict(org=org, N=N, T=T, r90=r90, r99=r99, rank_frac=r99 / N,
                   ac_med=float(np.median(ac)), ac_frac_smooth=float(np.mean(ac > 0.5)),
                   n_ion=n_ion, n_kegg=n_kegg, n_ref=n_ref, ref=ref_name)
    print(f"saved {out}")
    print(f"  {org:14s}: rank99={r99}/{N} ({100*r99/N:.0f}%) | "
          f"autocorr med={np.median(ac):.2f} smooth%={100*np.mean(ac>0.5):.0f} | "
          f"{n_ref}/{n_ion} in {ref_name}")
    return verdict


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ref_kegg = model_kegg_ids(os.path.join(ROOT, "papers/iJO1366.xml"))
    ref_name = "iJO1366"
    print("=== cross-organism amenability (common reference network: iJO1366) ===")
    res = []
    for i, (org, slug, ds, as_) in enumerate(ORGANISMS):
        t, C, idx2kegg = load_organism(wb, ds, as_)
        res.append(run_one(org, slug, "abcd"[i], t, C, idx2kegg, ref_kegg, ref_name))
    wb.close()
    print("\n=== generalisation summary ===")
    for r in res:
        print(f"  {r['org']:14s}: low-rank {r['rank_frac']*100:.0f}% | "
              f"dc/dt {'SMOOTH' if r['ac_med']>0.5 else 'NOISY'} (med {r['ac_med']:.2f}) | "
              f"coverage {100*r['n_ref']/r['n_ion']:.0f}%")
    return res


if __name__ == "__main__":
    main()
