#!/usr/bin/env python
"""
Real E. coli metabolomics — amenability probes (H5/H6/H7).

Question: is the GNN inverse approach even *amenable* to the real Link/Sauer
(nmeth.3584) E. coli real-time metabolomics, before attempting any training?

Recapitulates the analysis behind figures/metabolism/ecoli_amenability.png:
  H5 low-rank      : SVD of the (ion x time) concentration matrix -> rank@90/99%.
  H6 predictability: per-ion lag-1 autocorrelation of the traces (smooth signal
                     vs white noise) -> is dc/dt structured enough to learn?
  H7 network cover : fraction of measured KEGG-annotated ions that map onto a
                     real E. coli network (e_coli_core, iJO1366) -> can we build S?

Inputs (vendored, read-only):
  papers/nmeth3584/41592_2015_BFnmeth3584_MOESM197_ESM.xlsx  (sheets Ecoli1/2/3,
                                                              "Annotation Ecoli")
  papers/e_coli_core.xml , papers/iJO1366.xml  (KEGG-annotated SBML)
Output:
  figures/metabolism/ecoli_amenability.png  + printed verdict table.
Run:  python figures/ecoli_amenability.py
"""
import os, re, glob
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "papers/nmeth3584/41592_2015_BFnmeth3584_MOESM197_ESM.xlsx")
OUT  = os.path.join(ROOT, "figures/metabolism/ecoli_amenability.png")
os.makedirs(os.path.dirname(OUT), exist_ok=True)


def load_ecoli_replicates():
    """Return list of (name, t[T], C[N,T]) for Ecoli1/2/3, and index->KEGG set."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    reps = []
    for sheet in ("Ecoli1", "Ecoli2", "Ecoli3"):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # row 0: time header; row 1: time values; rows 2+: ion rows (index, v...)
        tvec = np.array([v for v in rows[1][1:] if v is not None], dtype=float)
        T = len(tvec)
        data = []
        for r in rows[2:]:
            if r[0] is None:
                continue
            vals = [np.nan if v is None else float(v) for v in r[1:1 + T]]
            if len(vals) == T:
                data.append(vals)
        C = np.array(data, dtype=float)  # (N, T)
        reps.append((sheet, tvec, C))
    # annotation: ion index -> set(KEGG)
    ann = wb["Annotation Ecoli"]
    idx2kegg = {}
    for row in list(ann.iter_rows(min_row=2, values_only=True)):
        idx, _mz, kegg = row[0], row[1], row[2]
        if idx is None:
            continue
        idx2kegg.setdefault(int(idx), set())
        if kegg:
            idx2kegg[int(idx)].add(str(kegg).strip())
    wb.close()
    return reps, idx2kegg


def model_kegg_ids(xml_path):
    """All distinct KEGG compound C-ids referenced in an SBML model."""
    if not os.path.exists(xml_path):
        return None
    txt = open(xml_path, "r", errors="ignore").read()
    # only count C-ids that appear in a kegg.compound annotation context when
    # possible; fall back to any C##### token (the models annotate via KEGG URLs)
    ids = set(re.findall(r"kegg\.compound[/:]?(C\d{5})", txt))
    if not ids:
        ids = set(re.findall(r"\bC\d{5}\b", txt))
    return ids


def zscore_rows(C):
    mu = np.nanmean(C, axis=1, keepdims=True)
    sd = np.nanstd(C, axis=1, keepdims=True)
    sd[sd == 0] = 1.0
    Z = (C - mu) / sd
    return np.nan_to_num(Z)


def svd_rank(Z):
    s = np.linalg.svd(Z, compute_uv=False)
    var = s ** 2
    cum = np.cumsum(var) / var.sum()
    r90 = int(np.searchsorted(cum, 0.90) + 1)
    r99 = int(np.searchsorted(cum, 0.99) + 1)
    return cum, r90, r99


def lag1_autocorr(C):
    """Per-ion lag-1 autocorrelation of the raw trace (NaN-safe)."""
    acs = []
    for row in C:
        x = row[~np.isnan(row)]
        if len(x) < 5 or np.std(x) == 0:
            continue
        x = x - x.mean()
        denom = np.sum(x * x)
        ac = np.sum(x[1:] * x[:-1]) / denom if denom > 0 else 0.0
        acs.append(ac)
    return np.array(acs)


def main():
    reps, idx2kegg = load_ecoli_replicates()
    name, t, C = reps[0]                      # Ecoli1 for the main panels
    N, T = C.shape
    Z = zscore_rows(C)
    cum, r90, r99 = svd_rank(Z)
    ac = lag1_autocorr(C)

    # H7 coverage
    core = model_kegg_ids(os.path.join(ROOT, "papers/e_coli_core.xml"))
    ijo  = model_kegg_ids(os.path.join(ROOT, "papers/iJO1366.xml"))
    ion_ids = sorted(idx2kegg)
    n_ion = len(ion_ids)
    n_with_kegg = sum(1 for i in ion_ids if idx2kegg[i])
    n_unambig   = sum(1 for i in ion_ids if len(idx2kegg[i]) == 1)
    def cover(model):
        if model is None:
            return None
        return sum(1 for i in ion_ids if idx2kegg[i] & model)
    n_core = cover(core)
    n_ijo  = cover(ijo)

    # ---- figure ----
    fig, ax = plt.subplots(2, 2, figsize=(12, 8))

    # (a) SVD low-rank
    ax[0, 0].plot(np.arange(1, len(cum) + 1), cum, "o-", ms=3, color="#2c3e50")
    ax[0, 0].axhline(0.99, ls="--", c="#e74c3c", alpha=.6)
    ax[0, 0].axhline(0.90, ls="--", c="#f39c12", alpha=.6)
    ax[0, 0].axvline(r99, ls=":", c="#e74c3c", alpha=.6)
    ax[0, 0].set_title(f"(a) H5 low-rank: SVD of C ({N} ions x {T} t)\n"
                       f"rank@90%={r90}, rank@99%={r99}  (synthetic ~35-47)")
    ax[0, 0].set_xlabel("singular component"); ax[0, 0].set_ylabel("cum. variance")
    ax[0, 0].set_xlim(0, min(60, len(cum)))

    # (b) autocorrelation distribution
    ax[0, 1].hist(ac, bins=30, color="#3498db", edgecolor="white")
    ax[0, 1].axvline(np.median(ac), c="#e74c3c", label=f"median={np.median(ac):.2f}")
    ax[0, 1].set_title("(b) H6 predictability:\nper-ion lag-1 autocorrelation "
                       "(1=smooth, 0=white noise)")
    ax[0, 1].set_xlabel("lag-1 autocorr"); ax[0, 1].set_ylabel("# ions"); ax[0, 1].legend()

    # (c) example traces (highest-variance ions, z-scored)
    order = np.argsort(-np.nanvar(C, axis=1))[:6]
    for j in order:
        ax[1, 0].plot(t, Z[j], lw=1)
    ax[1, 0].set_title("(c) 6 highest-variance ions (z-scored)\nreal-time dynamics")
    ax[1, 0].set_xlabel("time"); ax[1, 0].set_ylabel("z(intensity)")

    # (d) network coverage
    labels = ["ions", "with\nKEGG", "unambig.", "in\ne_coli_core", "in\niJO1366"]
    vals = [n_ion, n_with_kegg, n_unambig,
            n_core if n_core is not None else 0,
            n_ijo if n_ijo is not None else 0]
    bars = ax[1, 1].bar(labels, vals, color=["#95a5a6", "#3498db", "#2980b9",
                                             "#27ae60", "#16a085"])
    for b, v in zip(bars, vals):
        ax[1, 1].text(b.get_x() + b.get_width() / 2, v + 2, str(v),
                      ha="center", fontsize=9)
    ax[1, 1].set_title("(d) H7 network coverage of measured ions\n"
                       "(can we build S from a real E. coli model?)")
    ax[1, 1].set_ylabel("# ions")

    fig.suptitle("Real E. coli metabolomics (Link et al. 2015, nmeth.3584) — "
                 "amenability probes", fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.97])
    fig.savefig(OUT, dpi=130)
    print(f"saved {OUT}")

    # ---- verdict ----
    print("\n=== amenability verdict (Ecoli1) ===")
    print(f"H5 low-rank      : rank@90%={r90}, rank@99%={r99} of {N} ions "
          f"-> {'LOW-RANK like synthetic' if r99 < N/2 else 'NOT low-rank'}")
    print(f"H6 predictability: median lag-1 autocorr={np.median(ac):.2f}, "
          f"frac>0.5={np.mean(ac>0.5):.2f} "
          f"-> {'SMOOTH/structured' if np.median(ac)>0.5 else 'noisy'}")
    print(f"H7 coverage      : {n_ion} ions, {n_with_kegg} KEGG-annotated, "
          f"{n_unambig} unambiguous; in e_coli_core={n_core}, in iJO1366={n_ijo}")
    return dict(N=N, T=T, r90=r90, r99=r99, ac_med=float(np.median(ac)),
                n_ion=n_ion, n_kegg=n_with_kegg, n_core=n_core, n_ijo=n_ijo)


if __name__ == "__main__":
    main()
